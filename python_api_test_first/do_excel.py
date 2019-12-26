#!/usr/bin/env python
#-*- coding: utf-8 -*-
'''
@File  : do_excel.py
@Date  : 2019/12/25 0025 15:56
@Author: xibei
'''

from openpyxl import load_workbook
import pandas


class DoExcel:

    @staticmethod
    def do_excel(file_name, sheet_name):
        '''
        利用 openpyxl 进行读取数据
        :param file_name: 工作薄地址
        :param sheet_name: 表单名
        :return: 返回list包裹的字典数据
        '''
        wb = load_workbook(file_name)
        sheet = wb[sheet_name]
        test_case = []
        for i in range(2, sheet.max_row+1):
            single_case = {}
            single_case[sheet.cell(1, 1).value] = sheet.cell(i, 1).value
            single_case[sheet.cell(1, 2).value] = sheet.cell(i, 2).value
            single_case[sheet.cell(1, 3).value] = sheet.cell(i, 3).value
            single_case[sheet.cell(1, 4).value] = sheet.cell(i, 4).value
            single_case[sheet.cell(1, 5).value] = sheet.cell(i, 5).value
            single_case[sheet.cell(1, 6).value] = sheet.cell(i, 6).value

            test_case.append(single_case)

        return test_case

    @staticmethod
    def do_pandas(file_name, sheet_name):
        '''
        利用pandas读取数据
        :param file_name: 工作薄地址
        :param sheet_name: 表单名
        :return: 返回list包裹的字典数据
        '''

        sheet_data = pandas.read_excel(file_name, sheet_name=sheet_name)
        # 获取表头
        sheet_title = sheet_data.columns.values
        # 获取行索引
        row_index = sheet_data.index.values

        test_case = []
        for j in row_index:
            row_data = sheet_data.loc[j, sheet_title].to_dict()
            test_case.append(row_data)
        return test_case







if __name__ == '__main__':
    a = DoExcel()
    # for i in a.do_excel('test_case.xlsx', 'Sheet1'):
    #     print(i)
    #     print()

    for i in a.do_pandas('test_case.xlsx', 'Sheet1'):
        print(i)
        print()