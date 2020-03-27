#!/usr/bin/env python
#-*- coding: utf-8 -*-
'''
@File  : do_excel.py
@Date  : 2020/1/17 0017 10:42
@Author: xibei
'''


from openpyxl import load_workbook


class Case:

    def __init__(self):
        self.id = None
        self.url = None
        self.method = None
        self.title = None
        self.params = None
        self.expect = None


class DoExcel:

    def __init__(self, file_name):
        self.file_name = file_name
        try:
            self.wb = load_workbook(filename=self.file_name)
        except FileNotFoundError as e:
            print("{0} not found,pleas check file path".format(self.file_name))
            raise e

    def get_case(self, sheet_name=None):
        """
        读取测试用例
        :param sheet_name: 待获取的表单名
        :return: 返回获取的测试用例
        """
        if sheet_name:
            sheet = self.wb[sheet_name]
        else:
            sheet_name = self.wb.sheetnames[0]  # 如果未传入表单，默认读取第一个表单
            sheet = self.wb[sheet_name]
        max_row = sheet.max_row  # 获取最大行
        case_data = []  # 收集所有的用例
        for r in range(2, max_row+1):  # openpyxl 读取excel的索引从1开始，第一行为标题 故从第二行开始读取
            case = Case()  # 实例一个Case ，存储每一行的每条用例
            case.id = sheet.cell(row=r, column=1).value
            case.url = sheet.cell(row=r, column=2).value
            case.method = sheet.cell(row=r, column=3).value
            case.title = sheet.cell(row=r, column=4).value
            case.params = sheet.cell(row=r, column=5).value
            case.expect = sheet.cell(row=r, column=6).value
            case_data.append(case)
        return case_data

    def write_by_case_id(self, sheet_name, case_id, column, value):
        """
        Excel写入函数
        :param sheet_name: 表单名
        :param case_id: 用例id
        :param column: 写入的列
        :param value: 写入值
        :return:
        """
        sheet = self.wb[sheet_name]
        max_row = sheet.max_row  # 获取最大行
        for r in range(2, max_row+1):  # 遍历用例id 如果相同 写入
            # print("获取excel中id为{}".format(sheet.cell(r, 1).value))
            if sheet.cell(r, 1).value == case_id:
                sheet.cell(r, column).value = value
                self.wb.save(filename=self.file_name)
                break
            # else:
            #     # print("case_id 对比失败")
            #     print("获取excel中id为{}".format(sheet.cell(r,1).value))
            #     # print("输入的id为{}".format(case_id))
            #
            # print('*******'*10)







if __name__ == '__main__':

    from python_api_test_qianchendai.common.http_requets import HttpRequest
    import json
    from python_api_test_qianchendai.common.read_conf import ReadConf

    wb = DoExcel(r"C:\Users\Administrator\Desktop\cases .xlsx")
    sheet_name = 'login'
    case_data = wb.get_case(sheet_name)
    read_conf = ReadConf()  # 创建实例 读取测试环境配置文件中的url
    for case in case_data:
        url = read_conf.get_conf_str("api", "url")+case.url
        params = json.loads(case.params)  # excel 获取的数据都是str，故需转化成字典
        resp = HttpRequest(method=case.method, url=url, params=params)

        print("输入的case_id为{}".format(case.id))
        wb.write_by_case_id(sheet_name=sheet_name, case_id=case.id, column=7, value=resp.get_text())
        # print(case.expect)
        # print(type(case.expect))
        if resp.get_json() == json.loads(case.expect):
            wb.write_by_case_id(sheet_name=sheet_name, case_id=case.id, column=8, value="pass")
            print("pass")
        else:
            wb.write_by_case_id(sheet_name=sheet_name, case_id=case.id, column=8, value="false")
            print("false")


