#!/usr/bin/env python
#-*- coding: utf-8 -*-
'''
@File  : python_20191220_openpyxl模块.py
@Date  : 2019/12/20 0020 16:09
@Author: xibei
'''

''' openpyxl '''

# 1.第三方库，需下载安装：pip install openpyxl
# 2. openpyxl下的load_workbook模块 ，支持excel的读写
# 3.excel 格式需为 .xlsx 后缀结尾
# 4. excel中的行、列索引从1开始
# 5.在写入数据或保存excel时 需要关闭excel文件，否则会报错
# 6.使用：
#     1）：load_workbook(文件地址) 打卡工作薄
#     2）：work_book[表单名] 打开表单
#     3）：sheet.cell(行, 列).value  读取单元格数据
#     4）：sheet.cell(行, 列).value=值  赋值
#     5）：sheet.max_column，sheet.max_row 获取最大行列
#     6）：work_book.sheetnames 获取所有的表单
#     7）：work_book.save('文件名.xlsx') 保存



from openpyxl import load_workbook

# 读取数据

# 第一步： load_workbook(文件地址) 打卡工作薄
work_book = load_workbook(r'D:\test\python_12_xibei\python_api_test_first\test_case.xlsx')  # 返回工作薄

# 第二步： work_book[表单名] 打开表单
sheet = work_book['Sheet1']

# 第三步： sheet.cell(行, 列).value  读取单元格数据
# a = sheet.cell(1, 1).value
# print(a)

# 赋值 sheet.cell()
# sheet.cell(1, 1).value = 'case_id'
# work_book.save(r'D:\test\python_12_xibei\python_api_test_first\test_case.xlsx')  # 保存工作薄
# work_book.save(新的文件名) # 另存为

# 获取最大行列
# max_clo = sheet.max_column
# max_row = sheet.max_row
# print(max_row)
# print(max_clo)

# 获取所有的表单
# sheets = work_book.sheetnames  # 返回list的表单名
# print(sheets)

for i in sheet.rows():
    for j in i:
        print(j.value)
